# -*- coding: utf-8 -*-
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

path = r"D:\개인\04_Personal\02. Finance\2. Gift_Money\관리_경조사_선물_기록.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)


def yn(v):
    if v is None or v == "":
        return None
    s = str(v).strip().upper()
    if s == "Y":
        return True
    if s == "N":
        return False
    return None


def slugify(s):
    s = re.sub(r"[^\w가-힣]+", "-", str(s).strip(), flags=re.UNICODE)
    s = re.sub(r"-+", "-", s).strip("-").lower()
    return s[:48] or "entry"


def as_date(v):
    if v is None or v == "":
        return None, False
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10], False
    s = str(v).strip()
    if s in ("?", "？"):
        return None, True
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s, False
    return s, False


items = []

ws = wb["축의금 내역"]
rows = list(ws.iter_rows(values_only=True))[1:]
event = "결혼식"
for r in rows:
    if r[0]:
        event = str(r[0]).strip()
    name = (r[2] or "").strip() if r[2] else ""
    if not name:
        continue
    d, unk = as_date(r[1])
    amt = r[5]
    amount = int(amt) if isinstance(amt, (int, float)) else None
    base = f"{(d or 'undated')}-{slugify(name)}"
    slug = base
    n = 1
    while any(i["slug"] == slug for i in items):
        n += 1
        slug = f"{base}-{n}"
    entry = {
        "id": slug,
        "slug": slug,
        "kind": "congratulatory",
        "eventType": event,
        "date": d,
        "dateUnknown": True if unk else None,
        "name": name,
        "amount": amount,
        "invited": yn(r[3]),
        "attended": yn(r[4]),
        "note": (str(r[6]).strip() if r[6] else None) or None,
    }
    items.append({k: v for k, v in entry.items() if v is not None})

ws = wb["조의금 내역"]
rows = list(ws.iter_rows(values_only=True))[1:]
event = "장례식"
for r in rows:
    if r[0]:
        event = str(r[0]).strip()
    name = (r[3] or "").strip() if r[3] else ""
    if not name:
        continue
    relation = (str(r[1]).strip() if r[1] else None) or None
    d, unk = as_date(r[2])
    amt = r[4]
    amount = int(amt) if isinstance(amt, (int, float)) else None
    base = f"{(d or 'undated')}-{slugify(name)}-{slugify(relation or 'x')}"
    slug = base
    n = 1
    while any(i["slug"] == slug for i in items):
        n += 1
        slug = f"{base}-{n}"
    entry = {
        "id": slug,
        "slug": slug,
        "kind": "condolence",
        "eventType": event,
        "relation": relation,
        "date": d,
        "dateUnknown": True if unk else None,
        "name": name,
        "amount": amount,
    }
    items.append({k: v for k, v in entry.items() if v is not None})

items.sort(key=lambda it: it.get("date") or "0000-00-00", reverse=True)

out = Path("src/content/finance/occasions.json")
out.parent.mkdir(parents=True, exist_ok=True)
out.write_text(json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print("wrote", len(items), "to", out)
