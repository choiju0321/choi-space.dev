# -*- coding: utf-8 -*-
"""Duo 소개팅 프로필 엑셀 → JSON + private/media 사진 추출"""
from __future__ import annotations

import json
import re
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = Path(r"d:\개인\04_Personal\07. Activity\소개팅\소개팅_프로필_DB.xlsx")
OUT = ROOT / "src" / "content" / "personal" / "dating-profiles.json"
MEDIA_ROOT = ROOT / "private" / "media" / "personal" / "dating"


def parse_sheet(ws, sheet_name: str) -> dict:
    category = None
    by_cat: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        cat, key, val = row[0], row[1], row[2]
        if cat is not None and str(cat).strip():
            category = str(cat).strip()
        if key is None:
            continue
        key_s = str(key).strip()
        if val is None:
            value = None
        elif isinstance(val, float) and val.is_integer():
            value = str(int(val))
        elif isinstance(val, (int, float)):
            value = str(val)
        else:
            value = str(val).strip()
            if value in {"", "-"}:
                value = None
        by_cat.setdefault(category or "", []).append({"key": key_s, "value": value})

    def map_of(cat: str) -> dict[str, str | None]:
        return {f["key"]: f["value"] for f in by_cat.get(cat, [])}

    def first_text(cat: str, keys: tuple[str, ...]) -> str | None:
        for f in by_cat.get(cat, []):
            if f["key"] in keys and f["value"]:
                return f["value"]
        return None

    basic = map_of("기본정보")
    edu = [
        {"level": f["key"], "detail": f["value"]}
        for f in by_cat.get("학력사항", [])
        if f["value"]
    ]

    jobs: list[dict] = []
    cur: dict | None = None
    job_start_keys = {"현직", "재직", "회사", "전직"}
    job_field_map = {
        "부서": "department",
        "직위": "title",
        "직급": "title",
        "담당업무": "field",
        "직무분야": "field",
        "근무지": "location",
    }
    for f in by_cat.get("직장사항", []) + by_cat.get("직장정보", []):
        if f["key"] in job_start_keys:
            if cur:
                jobs.append(cur)
            cur = {
                "company": f["value"],
                "role": (
                    "current"
                    if f["key"] in {"현직", "재직", "회사"}
                    else "previous"
                ),
            }
        elif cur is not None:
            if f["key"] in job_field_map:
                cur[job_field_map[f["key"]]] = f["value"]
            else:
                cur.setdefault("extra", {})[f["key"]] = f["value"]
    if cur:
        jobs.append(cur)

    family = {k: v for k, v in map_of("가족사항").items() if v}

    intro = first_text("자기소개", ("내용",))
    ideal = (
        first_text("희망상대 스타일", ("내용",))
        or first_text("희망하는 스타일", ("내용",))
        or first_text("이상형 스타일", ("내용",))
    )

    manager_cat = (
        "매칭매니저 리뷰"
        if "매칭매니저 리뷰" in by_cat
        else "매칭매니저 의견"
    )
    manager_note = first_text(manager_cat, ("요약", "내용"))
    manager_name = first_text(manager_cat, ("이름",))
    manager_phone = first_text(manager_cat, ("연락처",))

    m = re.match(r"(\d{8})_?[Pp]rofile_?(\d+)", sheet_name)
    met_at = None
    idx = None
    if m:
        d = m.group(1)
        met_at = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
        idx = int(m.group(2))

    # 시트명이 20250722여도 실제 수령은 2026-07-22
    SHEET_MET_AT_OVERRIDE = {
        "20250722_profile_1": "2026-07-22",
        "20250722_profile_2": "2026-07-22",
    }
    if sheet_name in SHEET_MET_AT_OVERRIDE:
        met_at = SHEET_MET_AT_OVERRIDE[sheet_name]

    member_id = basic.get("이름(회원번호)")
    if member_id is not None:
        member_id = str(member_id).split(".")[0]

    slug = f"{met_at or 'unknown'}-{member_id or sheet_name}"

    birth_year = None
    by_label = basic.get("출생년도")
    if by_label:
        ym = re.search(r"(19|20)\d{2}", str(by_label))
        if ym:
            birth_year = int(ym.group(0))

    height_cm = None
    height = basic.get("키")
    if height:
        hm = re.search(r"(\d{2,3})", str(height))
        if hm:
            height_cm = int(hm.group(1))

    return {
        "id": slug,
        "slug": slug,
        "sourceSheet": sheet_name,
        "platform": "duo",
        "metAt": met_at,
        "batchIndex": idx,
        "memberId": member_id,
        "gender": basic.get("성별"),
        "birthYear": birth_year,
        "birthYearLabel": basic.get("출생년도"),
        "surname": basic.get("성씨"),
        "residence": basic.get("거주지"),
        "religion": basic.get("종교"),
        "height": basic.get("키"),
        "heightCm": height_cm,
        "hobby": basic.get("취미"),
        "education": edu,
        "jobs": jobs,
        "family": family,
        "intro": intro,
        "idealType": ideal,
        "managerNote": manager_note,
        "managerName": manager_name,
        "managerPhone": manager_phone,
        "contactName": None,
        "contactPhone": None,
        "photos": [],
        "status": "archived",
        "note": None,
    }


def extract_photos(ws, slug: str) -> list[str]:
    folder = MEDIA_ROOT / slug
    if folder.exists():
        shutil.rmtree(folder)
    folder.mkdir(parents=True, exist_ok=True)

    photos: list[str] = []
    images = list(getattr(ws, "_images", []) or [])
    for index, img in enumerate(images, start=1):
        data = img._data() if hasattr(img, "_data") else None
        if not data:
            continue
        # openpyxl path hint may be stale; sniff format
        ext = ".png"
        if data[:3] == b"\xff\xd8\xff":
            ext = ".jpg"
        elif data[:4] == b"RIFF":
            ext = ".webp"
        file_name = f"{index:02d}{ext}"
        (folder / file_name).write_bytes(data)
        photos.append(f"personal/dating/{slug}/{file_name}")
    return photos


def main() -> None:
    # data_only for values; keep drawings by loading again without data_only
    wb_data = openpyxl.load_workbook(SRC, data_only=True)
    wb_draw = openpyxl.load_workbook(SRC, data_only=False)

    MEDIA_ROOT.mkdir(parents=True, exist_ok=True)
    test_dir = MEDIA_ROOT / "_extract_test"
    if test_dir.exists():
        shutil.rmtree(test_dir)

    profiles = []
    for name in wb_data.sheetnames:
        profile = parse_sheet(wb_data[name], name)
        profile["photos"] = extract_photos(wb_draw[name], profile["slug"])
        profiles.append(profile)

    profiles.sort(
        key=lambda p: (p.get("metAt") or "", p.get("batchIndex") or 0),
        reverse=True,
    )
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(profiles, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"wrote {OUT} ({len(profiles)})")
    for p in profiles:
        print(
            p["slug"],
            f"photos={len(p['photos'])}",
            p.get("surname"),
            p.get("birthYearLabel"),
        )


if __name__ == "__main__":
    main()
